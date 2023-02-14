import tkinter as tk

# Create the main window
root = tk.Tk()
root.title("Sing-in Screen")
root.attributes("-fullscreen", True)


class Screen:
    def excel(self):
        import xlsxwriter
        workbook = xlsxwriter.Workbook("sales_data.xlsx")
        teste = 1 

        # Add a new worksheet to the file
        worksheet = workbook.add_worksheet()
        username = self.username_entry.get()
        # Write data to the worksheet
        worksheet.write("A1", username)
        worksheet.write("A2", teste)
        worksheet.write("A3", "Sales")
        worksheet.write("B1", 1000)
        worksheet.write("B2", "Apples")
        worksheet.write("B3", 500)
        worksheet.write("C1", "Oranges")
        worksheet.write("C2", 500)
        worksheet.write("C3", 500)

        # Save and close the file
        workbook.close()

    def show_excel(self):
        import openpyxl
        workbook = openpyxl.load_workbook('sales_data.xlsx')
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows():
            for cell in row:
                tk.Label(root, text=cell.value).pack()

    def clear(self):
        # Create the label and entry for the username
        self.username_label = tk.Label(root, text="Username:")
        self.username_label.pack()

        self.username_entry = tk.Entry(root)
        self.username_entry.pack()

        # Create the buttons
        self.generate_excel_button = tk.Button(root, text="Generate Excel", command=self.excel)
        self.generate_excel_button.pack()

        self.show_excel_button = tk.Button(root, text="Show Excel", command=self.show_excel) 
        self.show_excel_button.pack()

        self.quit_button = tk.Button(root, text="Quit", command=root.quit)
        self.quit_button.pack()

    def __init__(self, root):
        self.clear()

app = Screen(root)

# Start the window's event loop
root.mainloop()
  